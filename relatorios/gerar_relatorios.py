import pandas as pd
from openpyxl import Workbook
from openpyxl.styles. import Font, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

def atualizar_planilha(caminho_arquivo: str):
    """
    Atualiza uma planilha existente com novos dados e formatações.

    Args:
        caminho_arquivo: O caminho completo para o arquivo Excel.
    """
    print(f"Iniciando a atualização da planilha: {caminho_arquivo}")

    try:
        # Carregar a planilha existente
        df = pd.read_excel(caminho_arquivo, sheet_name='Dados Brutos')

        # --- Lógica de Atualização/Processamento ---
        #1. Adicionar uma coluna calculada (Margem de Lucro = Vendas - Custo)
        if 'Vendas' in df.columns and 'Custo' in df.columns:
            df['Margem de Lucro'] = df['Vendas'] - df['Custo']
        
        #2. Filtrar dados
        df_filtrado = df[df['Margem de Lucro'] > 0].copy()

        #3. Adicionar uma nova linha de resumo
        total_vendas = df['Vendas'].sum()
        novo_registro = pd.DataFrame([['Total', '', total_vendas, 0, 0]], columns=df.columns)
        df_final = pd.concat([df_filtrado, novo_registro], ignore_index=True)

        # --- Geração e Formatação do Relatório ---
        # Criar um novo workbook para o relatorio final
        wb = Workbook()
        ws = wb.active
        ws.title = "Relatório Semanal"

        # Inserir o DataFrame processado na planilha
        for r_idx, row in enumerate(dataframe_to_rows(df_final, header=True, index=False)):
            ws.append(row)

        # Aplicar formatações
        # Cabeçalho em negrito
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        # Formatação de moeda para colunas relevantes
        for col in ['C', 'D', 'E']:  # Supondo que essas colunas sejam monetárias
            for cell in ws[col][1:]:
                cell.number_format = 'R$ #,##0.00'
        
        # Salvar o novo relatório
        novo_caminho = caminho_arquivo.replace('.xlsx', '_relatorio_final.xlsx')
        wb.save(novo_caminho)
        print(f"Relatório gerado com sucesso: {novo_caminho}")

    except FileNotFoundError:
        print(f"Erro: Arquivo não encontrado no caminho: {caminho_arquivo}")
    except Exception as e:
        print(f"Ocorreu um erro durante a atualização: {e}")
    
def criar_planilha_exemplo(caminho_arquivo):
    """
    Cria um arquivo Excel de exemplo para test.
    """
    try:
        dados = {
            'Item': ['A', 'B', 'C', 'D', 'E'],
            'Região': ['Norte', 'Sul', 'Leste', 'Oeste', 'Norte'],
            'Vendas': [150.50, 230.75, 95.00, 450.20, 180.00],
            'Custo': [50.00, 100.50, 40.00, 200.00, 75.00]
        }
        df = pd.DataFrame(dados)

        writer = pd.ExcelWriter(caminho_arquivo, engine='openpyxl')
        df.to_excel(writer, sheet_name='Dados Brutos', index=False)
        writer.save()
        print(f"Arquivo de exemplo criado em: {caminho_arquivo}")
    except Exception as e:
        print(f"Ocorreu um erro ao criar o arquivo de exemplo: {e}")
